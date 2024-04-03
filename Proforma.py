""" Proforma
The goal of this program is to save accounts
receivable employees time. Many customers expect
some sort of PDF/file with pricing on it before
giving CC information for prepayment. The old
process involved a generic text email. This program
creates a pdf, saves it and also has the option
to send to anyone via user input. It also includes
the account manager.
"""
import os
import time
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
from datetime import datetime
from datetime import date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from itoolkit.transport import DatabaseTransport
import pyodbc
#imports
pb = '\n'
#this function runs our two needed queries to get all of the order infomation
def query(cntrl,username,password):
    global wb,items,options,control
    control = cntrl
    #excel
    wb = Workbook()
    wb.remove(wb['Sheet'])
    items = wb.create_sheet('Items')
    options = wb.create_sheet('Options')
    #query info
    conn = pyodbc.connect('REDACTED',uid=username,password=password)
    itransport = DatabaseTransport(conn)
    cursor = conn.cursor()
    #query one for the items and query two for the adders
    query = """

    SELECT                                                             
    T01.CONTRL, T01.PONUM, T01.ACCTNO, T04.ODCTR, T04.ITEMNO,
    T05.ITMDSC, T04.UORDER, T04.BEPRIC, T03.NAME, T03.POBOX, T03.STR1,
    T03.STR2, T03.CITY, T03.SORP, T03.ZIP, T03.CNTRY, T03.CTYCOD,
    T03.LPHON, T02.EMAILA, T03.ACBUD
        
    FROM DTABUD.OCNHDRPF T01
    INNER JOIN DTABUD.CSRREP2PF T02 ON T01.CSR = T02.REPNO
    INNER JOIN DTABUD.ACTFILPF T03 ON T01.ACCTNO = T03.ACCTNO
    INNER JOIN DTABUD.OCNDTLPF T04 ON T01.CONTRL = T04.CONTRL
    INNER JOIN DTABUD.OITEMSPF T05 ON T04.ITEMNO = T05.ITEMNO
    
    WHERE
    T01.CONTRL = """+control+"""
    AND T02.REPTYP = 'C'
    AND T04.CTRCTT <> 'X'
    
    """
    
    query2 = """

    SELECT
    T01.CONTRL, T01.OOCTR, T02.OTDSC, T01.OUNIT, T01.OEPRIC,
    T01.ADDSUB, T01.INVPRT  
                                          
    FROM DTABUD.OCNOPTPF T01
    LEFT OUTER JOIN DTABUD.OITOPTPF T02 ON T01.PRCOPT = T02.PRCOPT
    
    WHERE
    T01.CONTRL = """+control+"""                               
    AND T01.INVPRT = 'Y'                                      
    AND T01.OITYPE = T02.ITMNTP
    AND T01.OODLT <> 'X'
        
    """
    
    cursor = cursor.execute(query)
    currentrow = 2

    for row in cursor:
        final = str(row)
        final = [elem for elem in row]
        for info in range(0,len(final)):
            cell = items.cell(row=currentrow,column=info+1)
            cell.value = final[info]
        currentrow = currentrow + 1

    cursor = cursor.execute(query2)
    currentrow = 2

    for row in cursor:
        final = str(row)
        final = [elem for elem in row]
        for info in range(0,len(final)):
            cell = options.cell(row=currentrow,column=info+1)
            cell.value = final[info]
        currentrow = currentrow + 1
#both queries place information into an excel sheet(poor choice by me but it works at this time)
#setup basically just reads that information and creates variables for later use
def setup():
    global address,acct,po,csr,name,bud
    for j in range(2,items.max_column+1):
        cell = items.cell(row=2,column=j)
        while True:
            if cell.value == '':
                break
            elif type(cell.value) != str:
                break
            elif cell.value.endswith(' '):
                cell.value = cell.value[:-1]
            elif cell.value.startswith(' '):
                cell.value = cell.value[2:]
            else:
                break
    #order info
    po = items.cell(row=2,column=2).value
    acct = str(items.cell(row=2,column=3).value)
    name = items.cell(row=2,column=9).value
    csr = items.cell(row=2,column=19).value
    bud = items.cell(row=2,column=20).value
    
    #address items
    pobox = items.cell(row=2,column=10).value
    city = items.cell(row=2,column=13).value
    state = items.cell(row=2,column=14).value
    postal = items.cell(row=2,column=15).value
    country = items.cell(row=2,column=16).value
    street1 = items.cell(row=2,column=11).value
    street2 = items.cell(row=2,column=12).value
    #creating one street object
    if street2 != None:
        street = street1 + ' ' + street2
    else:
        street = street1
    if pobox != '':
        street = pobox + ' ' + street
    #putting all the info together
    phone = str(items.cell(row=2,column=17).value) + '-' + str(items.cell(row=2,column=18).value)[0:3]+'-'+str(items.cell(row=2,column=18).value)[3:7]
    csz = city + ', ' + state + ' ' + postal + ' - ' + country
    address = 'Acct: '+acct+pb+name+pb+street+pb+csz+pb+phone
    

def proforma(username):    
    ##### PDF #####
    #the main problem at the time of writing this is that it is limited to one page
    #I later created a somewhat similar program that generates invoices and that allows for multiple pages
    #however, at the time of writing this I would probably have to rewrite this program and thus far we have not run into that problem
    global pdf,today,email,ourPhone,ourName
    #basic business information based company division
    if bud == 1:
        mmmaddress = REDACTED
        logo = r'REDACTED.png'
        logoW = 90
        logoH = 9
        link = 'REDACTED'
        ourPhone = 'REDACTED'
        email = 'REDACTED'
        ourName = 'REDACTED'
        currency = 'USD'
    elif bud == 9:
        mmmaddress = REDACTED
        logo = r'REDACTED.png'
        logoW = 90
        logoH = 9
        link = 'REDACTED'
        ourPhone = 'REDACTED'
        email = 'REDACTED'
        ourName = 'REDACTED'
        currency = 'CAD'
    elif bud == 10:
        mmmaddress = REDACTED
        logo = r'REDACTED.png'
        logoW = 90
        logoH = 9
        link = 'REDACTED'
        ourPhone = 'REDACTED'
        email = 'REDACTED'
        ourName = 'REDACTED'
        currency = 'USD'
    itemstotal = 0
    optionstotal = 0
    
    #setup - non unique information
    afix = 3#due to border of the document this was implemented, pain...
    pdf = FPDF('P', 'mm', 'A4')
    pdf.add_page()
    pdf.add_link()
    today = date.today()
    today = str(today.strftime("%m/%d/%y"))
    pdf.set_fill_color(236, 236, 236)
    pdf.set_font('Arial','B', 20)
    pdf.set_text_color(255,0,0)
    pdf.set_xy(14+afix,17.5)
    pdf.image(logo,w=logoW,h=logoH,link=link)
    pdf.set_xy(7.5+afix,35)
    pdf.set_font('Arial','B', 25)
    pdf.set_text_color(90,90,90)
    pdf.cell(0,20,'Prepayment Estimate', border=0, align='C')
    pdf.set_xy(7.5+afix,44)
    pdf.set_font('Arial','I', 14)
    pdf.set_text_color(255,0,0)
    pdf.cell(0,20,'All pricing is shown in '+currency, border=0, align='C')
    
    #customer info - contact, address stuff like that
    pdf.set_font('Arial','', 10)
    pdf.set_text_color(0,0,0)
    pdf.set_xy(15+afix,60)
    pdf.multi_cell(85,5,address, border=1, align='L',fill=True)
    pdf.set_xy(105+afix,60)
    pdf.multi_cell(85,5,mmmaddress, border=1, align='L',fill=True)
    pdf.set_font('Arial','B', 10)
    pdf.set_text_color(50,50,50)
    pdf.set_xy(160+afix,10)
    pdf.cell(50,5,'Date: '+today, border=0, align='L')
    pdf.set_xy(160+afix,15)
    pdf.cell(50,5,'Acct: '+acct, border=0, align='L')
    pdf.set_xy(160+afix,20)
    pdf.cell(50,5,'Order #'+control, border=0, align='L')
    pdf.set_xy(160+afix,25)
    pdf.cell(50,5,'PO #'+po, border=0, align='L')
    
    #customer items - actual product
    pdf.set_font('Arial','B', 13)
    pdf.set_xy(15+afix,98)
    pdf.cell(10,5,'Items', border=0, align='L')
    pdf.set_font('Arial','B', 10)
    pdf.set_xy(15+afix,105)
    pdf.cell(10,5,'No.', border=1, align='C',fill=True)
    pdf.set_xy(25+afix,105)
    pdf.cell(100,5,'Item Description', border=1, align='C',fill=True)
    pdf.set_xy(125+afix,105)
    pdf.cell(20,5,'Qty', border=1, align='C',fill=True)
    pdf.set_xy(145+afix,105)
    pdf.cell(25,5,'Price per unit', border=1, align='C',fill=True)
    pdf.set_xy(170+afix,105)
    pdf.cell(20,5,'Total', border=1, align='C',fill=True)
    x = 0
    y = 0
    
    for i in range(2,items.max_row+1):
        pdf.set_xy(x+15+afix,y+110)
        pdf.cell(10,5,str(items.cell(row=i,column=4).value),border=1, align='L')
        pdf.set_xy(x+25+afix,y+110)
        pdf.cell(100,5,items.cell(row=i,column=6).value,border=1, align='L')
        pdf.set_xy(x+125+afix,y+110)
        pdf.cell(20,5,str(items.cell(row=i,column=7).value), border=1, align='C')
        pdf.set_xy(x+145+afix,y+110)
        pdf.cell(25,5,'$'+str(round(items.cell(row=i,column=8).value/items.cell(row=i,column=7).value,4)), border=1, align='C')
        pdf.set_xy(x+170+afix,y+110)
        pdf.cell(20,5,'$'+format(items.cell(row=i,column=8).value,','), border=1, align='C')
        y = y+5
        itemstotal = itemstotal + items.cell(row=i,column=8).value
        
    #upcharges/options
    pdf.set_font('Arial','B', 13)
    pdf.set_xy(15+afix,y+112)
    pdf.cell(10,5,'Adders', border=0, align='L')
    y = y+8
    pdf.set_font('Arial','B', 10)
    third = 'TBD'
    
    if options.cell(row=2,column=1).value != None:
        for i in range(2,options.max_row+1):
            if '3RD PARTY SHIPPING' in options.cell(row=i,column=3).value:
                third = '3rd Party'
        for i in range(2,options.max_row+1):
            if options.cell(row=i,column=6).value != 'A':
                options.cell(row=i,column=5).value = options.cell(row=i,column=5).value * -1
            pdf.set_xy(x+15+afix,y+110)
            pdf.cell(10,5,str(options.cell(row=i,column=2).value),border=1, align='L')
            pdf.set_xy(x+25+afix,y+110)
            pdf.cell(100,5,options.cell(row=i,column=3).value,border=1, align='L')
            pdf.set_xy(x+125+afix,y+110)
            pdf.cell(20,5,str(round(options.cell(row=i,column=4).value)), border=1, align='C')
            pdf.set_xy(x+145+afix,y+110)
            pdf.cell(25,5,'$'+str(round(options.cell(row=i,column=5).value/options.cell(row=i,column=4).value,4)), border=1, align='C')
            pdf.set_xy(x+170+afix,y+110)
            pdf.cell(20,5,'$'+format(options.cell(row=i,column=5).value,','), border=1, align='C')
            y = y+5
            optionstotal = optionstotal + options.cell(row=i,column=5).value
    #totals
    optionstotal = round(optionstotal,2)
    itemstotal = round(itemstotal,2)
    total = round(optionstotal+itemstotal,2)
    pdf.set_xy(x+15+afix,y+110)
    pdf.cell(110,5,'Shipping & Handling',border=1, align='C',fill=True)
    pdf.set_xy(x+125+afix,y+110)
    pdf.cell(65,5,third, border=1, align='C',fill=True)
    y = y+3

    pdf.set_font('Arial','B', 13)
    pdf.set_xy(125+afix,y+112)
    pdf.cell(65,10,'Totals', border=0, align='R')
    y = y+10
    pdf.set_font('Arial','B', 10)
    #final touches and finishing pdf save
    final = 'Items: $'+format(itemstotal,',')+pb+'Adders: $'+format(optionstotal,',')+pb+'Shipping & Handling: '+third
    disclaimer = '*Please note that the final value of your order will be determined when your order ships and bills. '
    disclaimer = disclaimer + 'Final cost will fluctuate due to shipping and our overrun(5%) and underrun(10%) policy.'+pb
    disclaimer = disclaimer + 'Any changes to your order will affect the final cost.'
    pdf.set_font('Arial','I', 10)
    pdf.set_xy(15+afix,y+110)
    pdf.multi_cell(110,5,disclaimer,border=0,align='L')
    pdf.set_font('Arial','B', 10)
    pdf.set_xy(125+afix,y+110)
    pdf.multi_cell(65,5,final,border=1,align='R',fill=True)
    y = y+15
    pdf.set_xy(125+afix,y+110)
    pdf.cell(65,5,'*Estimate: $'+format(total,','), border=1, align='R')
    pdf.set_xy(125+afix,y+115)
    pdf.set_font('Arial','I', 10)
    pdf.set_text_color(255,0,0)
    pdf.cell(65,5,'All pricing is shown in '+currency, border=0, align='R')
    y = y+5

    emailDisc = 'Your order is on hold for prepayment. Please call '+ourPhone+' with your credit card or ACH account information.'
    pdf.set_font('Arial','B', 14)
    pdf.set_text_color(255,0,0)
    pdf.set_xy(15+afix,y+160)
    pdf.multi_cell(175,5,emailDisc, border=0, align='C')
    pdf.set_title(str(control))
    pdf.set_author(ourName+' - User: '+username)
    try:
        pdf.output(r'REDACTED/burge/Generated Estimates/'+str(acct)+' - '+str(control)+'.pdf')
        #wb.save('test.xlsx')
        #place this code back in the case you want to see how
        #numbers are placed per column
    except Exception as e:
        print('File not updated. Estimate for "'+str(acct)+' - '+str(control)+'" may already be open.')
        time.sleep(2)
        


def sendEmail(requester):
    #send an email to customer and include AM and AR
    errortest = acct
    server = smtplib.SMTP(REDACTED)
    msg = MIMEMultipart()
    msg['From'] = email
    msg['To'] = requester
    msg['Cc'] = email+';'+csr

    subject = 'Prepayment Estimate for PO #'+po
    msg['Subject'] = subject
    body1 = today+pb+'Acct: '+str(acct)+' - '+name+pb+pb
    body2 = 'See the attached document for your prepayment estimate. This estimate is in regards to PO #'+ po+'.'+pb+pb
    body3 = 'Please note your order is on hold for prepayment. '
    body4 = 'Call '+ourPhone+' with your credit card or ACH account information.'+pb+pb
    body5 = 'If you have any questions or concerns, let us know.'+pb+pb+'Kind regards,'+pb+pb
    body6 = 'Accounts Receivable Department'+pb+ourName+pb+'REDACTED'+pb
    body7 = 'REDACTED'
    body = body1+body2+body3+body4+body5+body6+body7
    msg.attach(MIMEText(body, 'plain'))
    
    filename = str(acct)+' - '+str(control)+'.pdf'
    path = r'REDACTED\\burge\\Generated Estimates\\'
    
    with open(path+filename, 'rb') as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",f"attachment; filename = {filename}")
    msg.attach(part)

    server.send_message(msg)
    del msg
    print('Estimate for control #'+str(control)+ ' sent to '+requester)
    server.quit()
#basically runs all the functions, copies the file to clipbaord and displays the file to the user
def build(entryGet,username,password):
    query(entryGet,username,password)
    setup()
    proforma(username)
    os.startfile(r'REDACTED/burge/Generated Estimates/'+str(acct)+' - '+str(control)+'.pdf')
    output = "'REDACTED\\burge\\Generated Estimates\\"+str(acct)+" - "+str(control)+".pdf'"
    os.system('powershell Set-Clipboard -LiteralPath '+output)

#build(str(1910721),'REDACTED','REDACTED')
#sendEmail('REDACTED')
