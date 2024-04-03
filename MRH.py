""" Morning Report Helper
The goal of this program is to aid billing employees
in their daily process of running certain morning
reports. Many of the manual processes are slow and
were fairly susceptible to human error. This program
automates key presses as well as running certain queries
on behalf of the user.
"""
from itoolkit.transport import DatabaseTransport
import pyodbc
import os
import time
from openpyxl import Workbook, load_workbook
import keyboard
#imports
#finished goods report, originally its own standalone, it has since been adjusted to work with burge
def finishedGoods(month,day,year,username,password,beginning,end):
    print('Finished Goods...')
    wb = Workbook()
    blank = []
    reports = []
    #opens a text file with each of the queries and runs them, program then opens the excel files where info is stored
    queries = open(r'REDACTED/queries.txt','r')
    queries = queries.read()
    queries = queries.replace('1stcontrol',beginning)
    queries = queries.replace('lastcontrol',end)
    queries = queries.replace('\n','').split('|')
    for i in queries:
        item = i.split('!')
        blank.append(item)
    queries = blank
    for query in queries:
        ws = wb.create_sheet(query[0])
        if query[0] != 'oeday6':
            reports.append(query[0])
        conn =  pyodbc.connect('REDACTED',uid=username,password=password)
        itransport = DatabaseTransport(conn)
        cursor = conn.cursor()
        cursor = cursor.execute(query[1])
        currentrow = 2
        titles = [i[0] for i in cursor.description]
        for i in range(0,len(titles)):
            cell = ws.cell(row=1,column=i+1)
            cell.value = titles[i]
        for row in cursor:
            final = str(row)
            final = [elem for elem in row]
            for info in range(0,len(final)):
                cell = ws.cell(row=currentrow,column=info+1)
                cell.value = final[info]
            currentrow = currentrow + 1
    numbers = []
    days = []
    daylist = []
    oeday6s = 0
    oeday6m = 0
    oeday6l = 0
    for report in reports:
        ws = wb[report]
        count = 0
        for row in range(2,ws.max_row+1):
            cell = ws.cell(row=row,column=2)
            if cell.value is not None:
                count = count+1
        numbers.insert(100,[report,count])
    ws = wb['oeday6']
    for j in range(2,ws.max_row+1):
        cell = ws.cell(row=j,column=8)
        if int(cell.value) <= 25000:
            oeday6s = oeday6s + 1
        elif int(cell.value) <= 75000:
            oeday6m = oeday6m + 1
        elif int(cell.value) > 75000:
            oeday6l = oeday6l + 1
    numbers.insert(100,['oeday6',oeday6s,oeday6m,oeday6l])
    print('Finished goods for control range: '+beginning+'-'+end)
    for num in numbers: 
        print(num)
    wb.save(r'REDACTED.xlsx')
    os.startfile(r'REDACTED.xlsx')
    os.startfile(r'REDACTED.xlsx')    
    print("Press 'p' for PP macro, 'm' for 3M macro or 'x' to return to the main menu. Close the program to exit.")
    count = 0
    return numbers
#large order report, originally a standalone has been integrated into burge
def oedaily(month,day,year,username,password):
    print('OEDaily...')
    labels = ['None',
              'Bud#',
              'Europe',
              'Acct#',
              'Control#',
              'Counter',
              'Units Ordered',
              'Item#',
              'Sheet Count',
              '# Color',
              'Mainline',
              'CSR',
              'Field rep',
              'Trep']
    wb = Workbook()
    ws = wb.active
    #opens a text file with query information and runs the query, saves a running copy and permanent copy
    oedaily2 = open(r'REDACTED','r')
    oedaily2 = oedaily2.read()
    oedaily2 = oedaily2.replace('oedate',"'"+year+'-'+month+'-'+day+"'")
    conn =  pyodbc.connect('REDACTED',uid=username,password=password)
    itransport = DatabaseTransport(conn)
    cursor = conn.cursor()
    cursor = cursor.execute(oedaily2)
    currentrow = 2
    for i in range(1,14):
        cell = ws.cell(row=1,column=i)
        cell.value = labels[i]
    for row in cursor:
        final = str(row)
        final = [elem for elem in row]
        for info in range(0,len(final)):
            cell = ws.cell(row=currentrow,column=info+1)
            cell.value = final[info]
        currentrow = currentrow + 1
    dir_name = "REDACTED"
    test = os.listdir(dir_name)
    for item in test:
        if item.endswith('.xlsx'):
            os.remove(os.path.join(dir_name, item))
            
    wb.save(r'REDACTED.xlsx')
    wb.save(r'REDACTED.xlsx')
#tax query also integrated into burge from a standalone
def tax(month,day,year,username,password):
    #essentially shows new information each day to make sure all orders that should be taxed are
    print('Tax...')
    wb = Workbook()
    ws = wb.active
    taxorder = open(r'REDACTED/taxorder.txt','r')
    taxorder = taxorder.read()
    conn =  pyodbc.connect('REDACTED',uid=username,password=password)
    itransport = DatabaseTransport(conn)
    cursor = conn.cursor()
    cursor = cursor.execute(taxorder)
    currentrow = 2
    labels = ['Contrl','Acct #','Order Date']
    for i in range(0,3):
        cell = ws.cell(row=1,column=i+1)
        cell.value = labels[i]
    for row in cursor:
        final = [elem for elem in row]
        for info in range(0,len(final)):
            cell = ws.cell(row=currentrow,column=info+1)
            cell.value = final[info]
        currentrow = currentrow + 1
    ob = load_workbook(r'REDACTED/previous day.xlsx')
    osh = ob.active
    for i in range(1,osh.max_row+1):
        for j in range(1,4):
            osh.cell(row=i,column=j).value = ''
    for i in range(1,osh.max_row+1):
        osh.cell(row=i,column=1).value = osh.cell(row=i,column=6).value
        osh.cell(row=i,column=2).value = osh.cell(row=i,column=5).value
        osh.cell(row=i,column=3).value = osh.cell(row=i,column=4).value
    for i in range(1,osh.max_row+1):
        for j in range(4,7):
            osh.cell(row=i,column=j).value = ''
    for i in range(1,ws.max_row+1):
        osh.cell(row=i,column=4).value = ws.cell(row=i,column=1).value
        osh.cell(row=i,column=5).value = ws.cell(row=i,column=2).value
        osh.cell(row=i,column=6).value = ws.cell(row=i,column=3).value     
    ob.save(r'REDACTED/previous day.xlsx')
    os.startfile(r'REDACTED/previous day.xlsx')
#basic query that shows any orders that are frozen, integrated into burge
def freeze(month,day,year,username,password):
    #it is a double check for the accounting department to make sure orders frozen should be
    print('Freeze...')
    wb = Workbook()
    ws = wb.active
    messages_f = open(r'REDACTED/messages_f.txt','r')
    messages_f = messages_f.read()
    conn =  pyodbc.connect('REDACTED',uid=username,password=password)
    itransport = DatabaseTransport(conn)
    cursor = conn.cursor()
    cursor = cursor.execute(messages_f)
    currentrow = 2
    labels = ['Control', 'Acct #', 'DueDate']
    for i in range(0,3):
        cell = ws.cell(row=1,column=i+1)
        cell.value = labels[i]
    for row in cursor:
        final = str(row)
        final = [elem for elem in row]
        for info in range(0,len(final)):
            cell = ws.cell(row=currentrow,column=info+1)
            cell.value = final[info]
        currentrow = currentrow + 1
    wb.save(r'REDACTED/Freeze.xlsx')
    os.startfile(r'REDACTED/Freeze.xlsx')


